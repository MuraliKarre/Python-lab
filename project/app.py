import openpyxl as xl   
from openpyxl.chart import BarChart, Reference
# print(cell.value)
# print(sheet.max_column)
# print(sheet.max_row)

#def process_workbook(filename):
wb = xl.load_workbook('filename.xlsx')
sheet = wb ['Sheet1']

for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row, 3)
            corrected_price = cell.value * 0.9
            corrected_proce_cell = sheet.cell(row, 4)
            corrected_proce_cell.value = corrected_price

values = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,
                min_col=4,
                max_col=4)

chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, 'e2')

wb.save('filename2.xlsx')
